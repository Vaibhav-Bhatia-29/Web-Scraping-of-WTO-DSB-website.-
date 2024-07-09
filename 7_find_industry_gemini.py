import pandas as pd
import google.generativeai as genai
import time
from tenacity import retry, stop_after_attempt, wait_fixed
import openpyxl

# Load your Excel file
excel_file = 'wto_dispute_cases.xlsx'
df = pd.read_excel(excel_file)

# Configure the Gemini API
genai.configure(api_key='AIzaSyBBfR5o6xF3g1CcvxWKlAyUFX8s-09WiRQ')

# Initialize the model
model = genai.GenerativeModel('gemini-pro')

@retry(stop=stop_after_attempt(5), wait=wait_fixed(2))
def generate_content_with_retry(prompt):
    return model.generate_content(prompt)

def process_case(row):
    case_no = row['Case No.']
    title = row['Title']
    subjects = [row[f'Subject_{i}_Name'] for i in range(1, 6) if pd.notna(row[f'Subject_{i}_Name'])]
    subjects_str = ', '.join(subjects)
    conclusion = row['Conclusion']
    
    prompt = f"""
    Analyze the following WTO dispute case:
    
    Title: {title}
    Subjects: {subjects_str}
    Conclusion: {conclusion}
    
    Based on this information, classify this case into one of the following industries:
    - Agriculture
    - Infrastructure
    - Financial
    - Chemicals
    - Manufacturing
    - Telecomm
    - Healthcare
    - Energy
    - Textiles
    - Miscellaneous

    Provide only the industry name as your response, nothing else. Use Miscellaneous only if you cannot classify in extreme cases  
    """
    
    try:
        response = generate_content_with_retry(prompt)
        return response.text.strip() if hasattr(response, 'text') else f"Error: No text response for case {case_no}"
    except Exception as e:
        return f"Error processing case {case_no}: {str(e)}"

try:
    # Load the workbook and select the active sheet
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active

    # Find the column index for 'Industry'
    industry_col = None
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=col).value == 'Industry':
            industry_col = col
            break

    # If 'Industry' column doesn't exist, add it
    if industry_col is None:
        industry_col = sheet.max_column + 1
        sheet.cell(row=1, column=industry_col, value='Industry')

    for index, row in df.iterrows():
        result = process_case(row)
        
        if result.startswith("Error"):
            print(result)
            sheet.cell(row=index+2, column=industry_col, value="Error: Could not classify")
            continue
        
        # Write the industry classification to the Excel sheet
        sheet.cell(row=index+2, column=industry_col, value=result)
        
        # Save the workbook after each case
        workbook.save(excel_file)
        
        print(f"Processed case {index + 1} of {len(df)}")
        time.sleep(1)  # 1-second delay between cases

    print(f"\nClassification complete. Results saved to {excel_file}")
except Exception as e:
    print(f"An error occurred: {str(e)}")
    print("Stopping the code execution.")
finally:
    # Make sure to close the workbook
    if 'workbook' in locals():
        workbook.close()