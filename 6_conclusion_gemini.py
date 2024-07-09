import pandas as pd
import google.generativeai as genai
import time
from tenacity import retry, stop_after_attempt, wait_fixed
import openpyxl

# Load your Excel file
excel_file = 'wto_dispute_cases.xlsx'
df = pd.read_excel(excel_file)

# Configure the Gemini API
genai.configure(api_key='AIzaSyBBfR5o6xF3g1CcvxWKlAyUFX8s-09WiRQ')

# Initialize the model
model = genai.GenerativeModel('gemini-pro')

@retry(stop=stop_after_attempt(5), wait=wait_fixed(2))
def generate_content_with_retry(prompt):
    return model.generate_content(prompt)

def process_case(row):
    case_no = row['Case No.']
    title = row['Title']
    subjects = row['Subject_1_Name']
    summary = row['Summary']
    
    prompt = f"""
    Analyze the following WTO dispute case:
    
    Title: {title}
    Subjects: {subjects}
    Summary: {summary}
    
    Please provide:
    1. A brief conclusion about this dispute with help of the summary (80 words).

    Format your response as follows:
    Conclusion: [Your conclusion here]
    """
    
    try:
        response = generate_content_with_retry(prompt)
        return response.text if hasattr(response, 'text') else f"Error: No text response for case {case_no}"
    except Exception as e:
        return f"Error processing case {case_no}: {str(e)}"

try:
    # Load the workbook and select the active sheet
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active

    # Find the column index for 'Conclusion'
    conclusion_col = None
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=col).value == 'Conclusion':
            conclusion_col = col
            break

    # If 'Conclusion' column doesn't exist, add it
    if conclusion_col is None:
        conclusion_col = sheet.max_column + 1
        sheet.cell(row=1, column=conclusion_col, value='Conclusion')

    for index in range(0, 5):  
        row = df.iloc[index]
        result = process_case(row)
        
        if result.startswith("Error"):
            print(result)
            sheet.cell(row=index+2, column=conclusion_col, value="Error: Could not process")
            continue  # Continue to the next case instead of stopping
        
        parts = result.split('\n')
        conclusion = parts[0].replace('Conclusion: ', '') if len(parts) > 0 else ''
        
        # Write the conclusion to the Excel sheet
        sheet.cell(row=index+2, column=conclusion_col, value=conclusion)
        
        # Save the workbook after each case
        workbook.save(excel_file)
        
        print(f"Processed case {index + 1} of {len(df)}")
        time.sleep(0.5)  # 0.5-second delay between cases

    print(f"\nAnalysis complete. Results saved to {excel_file}")
except Exception as e:
    print(f"An error occurred: {str(e)}")
    print("Stopping the code execution.")
finally:
    # Make sure to close the workbook
    if 'workbook' in locals():
        workbook.close()